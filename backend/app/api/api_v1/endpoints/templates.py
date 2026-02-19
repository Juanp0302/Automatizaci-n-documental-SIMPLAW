from typing import Any, List, Optional

from fastapi import APIRouter, Depends, HTTPException, File, UploadFile, Form
from sqlalchemy.orm import Session
import shutil
import os

from app import crud, models, schemas
from app.api import deps
from app.core.config import settings

router = APIRouter()


@router.get("/", response_model=List[schemas.Template])
def read_templates(
    db: Session = Depends(deps.get_db),
    skip: int = 0,
    limit: int = 100,
    current_user: models.User = Depends(deps.get_current_active_user),
) -> Any:
    """
    Retrieve templates.
    """
    if current_user.is_superuser:
        templates = crud.template.get_multi(db, skip=skip, limit=limit)
    else:
        templates = crud.template.get_multi_by_owner(
            db=db, owner_id=current_user.id, skip=skip, limit=limit
        )
    return templates


@router.post("/", response_model=schemas.Template)
def create_template(
    *,
    db: Session = Depends(deps.get_db),
    title: str = Form(...),
    description: Optional[str] = Form(None),
    file: UploadFile = File(...),
    current_user: models.User = Depends(deps.get_current_active_user),
) -> Any:
    """
    Create new template.
    """
    file_location = os.path.join(settings.TEMPLATES_DIR, file.filename)
    with open(file_location, "wb+") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    template_in = schemas.TemplateCreate(title=title, description=description, file_path=file_location)
    template = crud.template.create_with_owner(db, obj_in=template_in, owner_id=current_user.id)
    return template




@router.get("/{id}", response_model=schemas.Template)
def read_template(
    *,
    db: Session = Depends(deps.get_db),
    id: int,
    current_user: models.User = Depends(deps.get_current_active_user),
) -> Any:
    """
    Get template by ID.
    """
    template = crud.template.get(db, id=id)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
        
    if not current_user.is_superuser and (template.owner_id != current_user.id):
        raise HTTPException(status_code=400, detail="Not enough permissions")
        
    return template


@router.delete("/{id}", response_model=schemas.Template)
def delete_template(
    *,
    db: Session = Depends(deps.get_db),
    id: int,
    current_user: models.User = Depends(deps.get_current_active_user),
) -> Any:
    """
    Delete a template.
    """
    template = crud.template.get(db, id=id)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    if not current_user.is_superuser and (template.owner_id != current_user.id):
        raise HTTPException(status_code=400, detail="Not enough permissions")
        
    template = crud.template.remove(db, id=id)
    return template


@router.get("/{id}/variables", response_model=List[str])
def get_template_variables(
    *,
    db: Session = Depends(deps.get_db),
    id: int,
    current_user: models.User = Depends(deps.get_current_active_user),
) -> Any:
    """
    Get variables from a template.
    """
    template = crud.template.get(db, id=id)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    if not os.path.exists(template.file_path):
        # Return empty list if file is missing, or raise error?
        # Let's return empty to avoid breaking UI, but log it
        print(f"Warning: Template file not found at {template.file_path}")
        return []

    from app.utils import extract_variables_from_docx
    variables = extract_variables_from_docx(template.file_path)
    return variables


@router.put("/{id}", response_model=schemas.Template)
def update_template(
    *,
    db: Session = Depends(deps.get_db),
    id: int,
    template_in: schemas.TemplateUpdate,
    current_user: models.User = Depends(deps.get_current_active_user),
) -> Any:
    """
    Update template title/description.
    """
    template = crud.template.get(db, id=id)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
        
    if not current_user.is_superuser and (template.owner_id != current_user.id):
        raise HTTPException(status_code=400, detail="Not enough permissions")
        
    template = crud.template.update(db, db_obj=template, obj_in=template_in)
    return template


@router.get("/{id}/batch-template")
def get_batch_template(
    *,
    db: Session = Depends(deps.get_db),
    id: int,
    current_user: models.User = Depends(deps.get_current_active_user),
) -> Any:
    """
    Generate an Excel template for batch document generation.
    """
    template = crud.template.get(db, id=id)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    if not os.path.exists(template.file_path):
        raise HTTPException(status_code=404, detail="Template file not found on disk")

    from app.utils import extract_variables_from_docx
    variables = extract_variables_from_docx(template.file_path)
    
    # Create Excel
    import openpyxl
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Input"
    
    # Headers
    # Add standard headers if needed (e.g. "output_filename")
    headers = ["output_filename"] + variables
    
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header
        ws.column_dimensions[col_letter].width = 20
        
    # Add detailed instructions or metadata if needed
    
    # Save to buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    safe_title = "".join([c for c in template.title if c.isalnum() or c in (' ', '-', '_')]).strip()
    filename = f"Batch_Template_{safe_title}.xlsx"
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


from app import schemas

@router.post("/{id}/batch-generate", response_model=dict)
def batch_generate_documents(
    *,
    db: Session = Depends(deps.get_db),
    id: int,
    file: UploadFile = File(...),
    current_user: models.User = Depends(deps.get_current_active_user),
) -> Any:
    """
    Generate multiple documents from an uploaded Excel file.
    """
    template = crud.template.get(db, id=id)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
        
    if not os.path.exists(template.file_path):
        raise HTTPException(status_code=404, detail="Template file not found on disk")
        
    import openpyxl
    from docxtpl import DocxTemplate
    from datetime import datetime
    
    try:
        wb = openpyxl.load_workbook(file.file)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        
        # Identify variable columns
        # We expect headers to match variable names. "output_filename" is special.
        
        results = {
            "total": 0,
            "success": 0,
            "failed": 0,
            "generated_docs": [],
            "errors": []
        }
        
        output_dir = settings.GENERATED_DOCS_DIR
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row): continue # Skip empty rows
            
            results["total"] += 1
            
            row_data = dict(zip(headers, row))
            
            # Extract context variables
            context = {k: v for k, v in row_data.items() if k != "output_filename" and v is not None}
            
            # Determine filename
            custom_filename = row_data.get("output_filename")
            if not custom_filename:
                safe_title = "".join([c for c in template.title if c.isalnum() or c in (' ', '-', '_')]).strip()
                custom_filename = f"{safe_title}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{row_idx}"
            
            safe_filename = "".join([c for c in str(custom_filename) if c.isalnum() or c in (' ', '-', '_', '.')]).strip()
            if not safe_filename.endswith(".docx"):
                safe_filename += ".docx"
                
            full_filename = f"{current_user.id}_{id}_{safe_filename}"
            file_path = os.path.join(output_dir, full_filename)
            
            try:
                # 1. Render Document
                doc = DocxTemplate(template.file_path)
                
                # Default context additions
                if "date" not in context:
                     from datetime import timezone, timedelta
                     bogota_tz = timezone(timedelta(hours=-5))
                     context["date"] = datetime.now(bogota_tz).strftime("%Y-%m-%d")
                     
                doc.render(context)
                doc.save(file_path)
                
                # 2. Save to DB
                db_doc = models.Document(
                    title=str(custom_filename).replace(".docx", ""),
                    template_id=id,
                    user_id=current_user.id,
                    generated_file_path=file_path
                )
                db.add(db_doc)
                db.commit()
                db.refresh(db_doc)
                
                results["success"] += 1
                results["generated_docs"].append({"id": db_doc.id, "title": db_doc.title})
                
            except Exception as e:
                import traceback
                traceback.print_exc()
                results["failed"] += 1
                results["errors"].append(f"Row {row_idx}: {str(e)}")
                
        return results
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing Excel file: {str(e)}")
